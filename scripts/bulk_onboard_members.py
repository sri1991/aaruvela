"""
Bulk onboard members from 'Members List.xlsx' via the admin API.
Reads both sheets (Normal, Permanent) and calls POST /admin/create-member for each.

Usage:
    python scripts/bulk_onboard_members.py

Requires: openpyxl, requests
    pip install openpyxl requests
"""

import sys
import re
import openpyxl
import requests
from datetime import datetime

BASE_URL = "http://localhost:8000"
ADMIN_PHONE = "1112223333"
ADMIN_PIN = "1234"
EXCEL_PATH = "Members List.xlsx"

# ── helpers ───────────────────────────────────────────────────────────────────

def clean_name(name):
    if not name:
        return None
    return re.sub(r"[`']+", "", str(name)).strip()

def clean_phone(phone):
    if not phone:
        return None
    digits = re.sub(r"\D", "", str(phone))
    # Take last 10 digits
    digits = digits[-10:] if len(digits) >= 10 else digits
    return digits if len(digits) == 10 else None

def get_admin_token():
    resp = requests.post(f"{BASE_URL}/auth/verify-pin", json={
        "identifier": ADMIN_PHONE,
        "pin": ADMIN_PIN
    })
    if resp.status_code != 200:
        print(f"  ✗ Login failed: {resp.text}")
        sys.exit(1)
    token = resp.json().get("access_token")
    print(f"  ✓ Admin login successful\n")
    return token

def create_member(token, payload):
    resp = requests.post(
        f"{BASE_URL}/admin/create-member",
        json=payload,
        headers={"Authorization": f"Bearer {token}"}
    )
    return resp.status_code, resp.json()

# ── read sheets ───────────────────────────────────────────────────────────────

def read_normal_sheet(ws):
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, guardian, gender, gotram, dob, address, phone, occupation, _ = row
        name = clean_name(name)
        phone = clean_phone(phone)
        if not name or not phone:
            continue
        members.append({
            "full_name": name,
            "phone": phone,
            "role": "NORMAL",
            "zonal_committee": "",
            "regional_committee": "",
        })
    return members

def read_permanent_sheet(ws):
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, guardian, gender, gotram, dob, address, phone, email, occupation, _ = row
        name = clean_name(name)
        phone = clean_phone(phone)
        if not name or not phone:
            continue
        members.append({
            "full_name": name,
            "phone": phone,
            "role": "PERMANENT",
            "zonal_committee": "",
            "regional_committee": "",
        })
    return members

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Parishat — Bulk Member Onboarding")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Load Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        print(f"\n✗ File not found: {EXCEL_PATH}")
        print("  Run this script from the project root directory.")
        sys.exit(1)

    normal_members   = read_normal_sheet(wb["Normal"])
    permanent_members = read_permanent_sheet(wb["Permanent"])
    all_members = permanent_members + normal_members

    print(f"\n  Found {len(permanent_members)} Permanent + {len(normal_members)} Normal = {len(all_members)} members\n")

    # Login
    print("  Authenticating as admin...")
    token = get_admin_token()

    # Track results
    created   = []
    updated   = []
    skipped   = []
    seen_phones = {}

    for i, member in enumerate(all_members, 1):
        phone = member["phone"]
        name  = member["full_name"]
        role  = member["role"]

        # Warn about shared phones (family members in same household)
        if phone in seen_phones:
            print(f"  [{i:>3}] ⚠  SKIP  {name} ({role}) — phone {phone} already used by {seen_phones[phone]}")
            skipped.append({"name": name, "phone": phone, "reason": f"Phone shared with {seen_phones[phone]}"})
            continue

        seen_phones[phone] = name
        status_code, resp = create_member(token, member)

        if status_code == 200:
            mid = resp.get("member_id", "—")
            print(f"  [{i:>3}] ✓ Created  {name} ({role}) → {mid}")
            created.append({"name": name, "phone": phone, "role": role, "member_id": mid})
        elif status_code == 409 or "already exists" in str(resp).lower():
            print(f"  [{i:>3}] ~  Exists   {name} ({role})")
            updated.append({"name": name, "phone": phone, "role": role})
        else:
            detail = resp.get("detail", resp)
            print(f"  [{i:>3}] ✗ Failed   {name} ({role}) — {detail}")
            skipped.append({"name": name, "phone": phone, "reason": str(detail)})

    # Summary
    print("\n" + "=" * 60)
    print(f"  SUMMARY")
    print(f"  Created  : {len(created)}")
    print(f"  Existing : {len(updated)}")
    print(f"  Skipped  : {len(skipped)}")
    print("=" * 60)

    if skipped:
        print("\n  Skipped details:")
        for s in skipped:
            print(f"    • {s['name']} ({s['phone']}): {s['reason']}")

    print("\n  Done.\n")

if __name__ == "__main__":
    main()
